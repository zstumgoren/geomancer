from flask import current_app
from pickle import loads, dumps
from redis import Redis
from uuid import uuid4
import sys
import os
from cStringIO import StringIO
from csvkit.unicsv import UnicodeCSVReader, UnicodeCSVWriter
from geo.utils.lookups import ACS_DATA_TYPES, GEO_TYPES
from geo.utils.census_reporter import CensusReporter, CensusReporterError
from geo.app_config import RESULT_FOLDER
from datetime import datetime
import xlwt
from openpyxl import Workbook
from openpyxl.cell import get_column_letter

redis = Redis()

class DelayedResult(object):
    def __init__(self, key):
        self.key = key
        self._rv = None

    @property
    def return_value(self):
        if self._rv is None:
            rv = redis.get(self.key)
            if rv is not None:
                self._rv = loads(rv)
        return self._rv
    
def queuefunc(f):
    def delay(*args, **kwargs):
        qkey = current_app.config['REDIS_QUEUE_KEY']
        key = '%s:result:%s' % (qkey, str(uuid4()))
        s = dumps((f, key, args, kwargs))
        redis.rpush(current_app.config['REDIS_QUEUE_KEY'], s)
        return DelayedResult(key)
    f.delay = delay
    return f

@queuefunc
def do_the_work(file_contents, field_defs, filename):
    """
      field_defs looks like:
      {
        10: {
          'type': 'city_state', 
          'append_columns': ['total_population', 'median_age']
        }
      }

      file_contents is a string containing the contents of the uploaded file.
    """
    contents = StringIO(file_contents)
    reader = UnicodeCSVReader(contents)
    header = reader.next()
    c = CensusReporter()
    result = None
    geo_ids = set()
    table_ids = set()
    geoid_mapper = {}
    output_filepath = 'output.csv'

    for row_idx, row in enumerate(reader):
        col_idxs = [int(k) for k in field_defs.keys()]
        for idx in col_idxs:
            val = row[idx]
            geo_type = field_defs[idx]['type']
            for column in field_defs[idx]['append_columns']:
                table_ids.add(ACS_DATA_TYPES[column]['table_id'])
            sumlevs = [g['acs_sumlev'] for g in GEO_TYPES if g['name'] == geo_type]
            try:
                if val and sumlevs:
                    geoid_search = c.geo_search(val, sumlevs=sumlevs)
                else:
                    continue
            except CensusReporterError, e:
                return e.message
            try:
                row_geoid = geoid_search['results'][0]['full_geoid']
                geo_ids.add(row_geoid)
            except IndexError:
                continue
            try:
                geoid_mapper[row_geoid].append(row_idx)
            except KeyError:
                geoid_mapper[row_geoid] = [row_idx]
    if geo_ids:
        try:
            data = c.data_show(geo_ids=list(geo_ids), table_ids=list(table_ids))
        except CensusReporterError, e:
            raise e
        header = data['header']
        contents.seek(0)
        all_rows = list(reader)
        included_idxs = set()
        header_row = all_rows.pop(0)
        output = []
        for col in header:
            header_row.append(col)
        output.append(header_row)
        for geoid, row_ids in geoid_mapper.items():
            for row_id in row_ids:
                included_idxs.add(row_id)
                row = all_rows[row_id]
                row.extend(data[geoid])
                output.append(row)
        all_row_idxs = set(list(range(len(all_rows))))
        missing_rows = all_row_idxs.difference(included_idxs)
        for idx in missing_rows:
            row = all_rows[idx]
            row.extend(['' for i in header])
            output.append(row)
        name, ext = os.path.splitext(filename)
        fname = '%s_%s%s' % (name, datetime.now().isoformat(), ext)
        fpath = '%s/%s' % (RESULT_FOLDER, fname)
        if ext == '.xlsx':
            writeXLSX(fpath, output)
        elif ext == '.xls':
            writeXLS(fpath, output)
        else:
            writeCSV(fpath, output)
        
        return '/download/%s' % fname
    else:
        raise CensusReporterError('No geographies matched')

def writeXLS(fpath, output):
    with open(fpath, 'wb') as f:
        workbook = xlwt.Workbook(encoding='utf-8')
        sheet = workbook.add_sheet('Geomancer Output')
        for r, row in enumerate(output):
            for c, col in enumerate(output[0]):
                sheet.write(r, c, row[c])
        workbook.save(fpath)

def writeXLSX(fpath, output):
    with open(fpath, 'wb') as f:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Geomancer Output'
        numcols = len(output[0])
        for r, row in enumerate(output):
            sheet.append([row[col_idx] for col_idx in range(numcols)])
        workbook.save(fpath)

def writeCSV(fpath, output):
    with open(fpath, 'wb') as f:
        writer = UnicodeCSVWriter(f)
        writer.writerows(output)


def queue_daemon(app, rv_ttl=500):
    while 1:
        msg = redis.blpop(app.config['REDIS_QUEUE_KEY'])
        func, key, args, kwargs = loads(msg[1])
        try:
            rv = func(*args, **kwargs)
            rv = {'status': 'ok', 'result': rv}
        except Exception, e:
            rv = {'status': 'error', 'result': e.message}
        if rv is not None:
            redis.set(key, dumps(rv))
            redis.expire(key, rv_ttl)